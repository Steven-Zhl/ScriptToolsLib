from csv import writer as csv_writer
from os import path as os_path
from os import startfile as os_startfile
from re import findall as re_findall
from re import match as re_match
from re import sub as re_sub

from openpyxl import load_workbook as load_excel


class Course:
    """
    这是一个用于存放原始课程信息的类，并提供适配各种软件单元格信息的接口
    不过目前只有HWCalender
    """

    def __init__(self):
        self.ID = ""  # 这门课程在云南大学中的ID，格式为"YN"+8位数字
        self.name = ""  # 课程名
        self.index = ""  # 本课程在同课程中的序号
        self.teacher = ""  # 教师
        self.weekIndex = ""  # 日期，XX-XX周
        self.week = ""  # 星期，星期X
        self.time = ""  # 时间，XX-XX节
        self.location = ""  # 上课位置，字符串格式记录
        self.QQ_Group = ""  # QQ群号，有的课会有这一项
        self.row = 0  # 这门课所在行
        self.column = 0  # 这门课所在列

        self.showID = False  # 是否显示ID
        self.showIndex = False  # 是否显示序号

    def HWCalenderMessage(self):
        """
        返回适用于华为日历的单个单元格中的内容
        """
        courseName = self.name  # 第一行，课程名称
        if self.showID:
            courseName = self.ID + "-" + courseName
        if self.showIndex:
            courseName = courseName + self.index
        weekAndTime = ""  # 第二行，周数节数
        if "周(单)" or "周(双)" in self.weekIndex:
            weekIndex = self.weekIndex.replace(
                "周(单)", "(单)周").replace("周(双)", "(双)周")
        else:
            weekIndex = self.weekIndex
        weekAndTime = weekIndex + self.time
        classroom = self.location  # 第三行，教室
        teacher = self.teacher  # 第四行，教师

        return "\n".join([courseName, weekAndTime, classroom, teacher])


class Apps:
    """
    这个类提供一些对于各种App都应有的基本操作
    """

    def __init__(self, AppName, sourcePath, time_1Week_Mon, saveDir):
        self.AppName = AppName  # 这个拿来确定导出时的文件名
        self.sourcePath = sourcePath  # 原xlsx文件路径
        self.time_1Week_Mon = time_1Week_Mon  # 第一周周一的日期，格式为20XX.XX.XX
        self.saveDir = saveDir  # 导出文件的文件夹路径
        self.Courses = []  # 原始的各种课程类信息，就是Course类的列表
        self.Courses_text = []  # 5×7的课程表格式
        self._readMessage()  # 读取文件
        self._convertClass2Text()  # 将读取到的信息转换为5×7的文本存储

    # 读取教务系统导出的课程表信息，课程信息使用Course类的列表存储
    def _readMessage(self):
        """
        读取教务系统导出的课程表信息，课程信息使用Course类的列表存储
        """

        sheet = load_excel(self.sourcePath).active  # 直接读取目标表单
        # 确认课程信息在excel中的位置
        # 课程信息在excel中实际存放的位置区间，用这个规范读取时的范围
        rowRange, columnRange = [7, 19], [2, 8]
        rowPhase, columnPhase = 4, 1  # 行、列的偏差值，用excel中的行列减去这个得到的相对行列从1开始，其实也就是正则化
        for i in range(sheet.max_row + 1, 1, -1):  # 确定行的范围
            if sheet.cell(i, 1).value == "9-10节":
                rowRange = [7, i + 2]
                break

        for row in range(rowRange[0], rowRange[1] + 1, 3):
            thisRowCourses = []  # 存放这一行的所有课程
            for column in range(columnRange[0], columnRange[1] + 1):
                content = sheet.cell(row=row, column=column).value
                if (content is not None) and (content != "") and (content != " "):
                    thisCourse = Course()
                    thisCourse.row, thisCourse.column = int(
                        (row - rowPhase) / 3), column - columnPhase  # 由于一节课占三行，所以/3
                    contents = content.split("\n")  # 划分一下数据，以便于进行进一步细分
                    # 以下是为类的属性赋值
                    # 第一行，课程相关信息
                    courseMessage = contents[0]
                    thisCourse.ID = courseMessage.split("-")[0]
                    thisCourse.name = re_sub(
                        "\[\d{2}]", "", courseMessage.split("-")[1])
                    thisCourse.index = re_findall(
                        "\[\d{2}]", courseMessage.split("-")[1])[0]
                    # 第二行，教师名
                    thisCourse.teacher = contents[1]
                    # 第三行，日期与位置
                    dateAndLocation = contents[2].split(",")
                    thisCourse.weekIndex, thisCourse.week = dateAndLocation[0], dateAndLocation[1]
                    thisCourse.time = dateAndLocation[2]
                    thisCourse.location = dateAndLocation[3]
                    # 第四行，QQ群信息（如果有的话）
                    thisCourse.QQ_Group = contents[4] if len(
                        contents) == 5 else None
                    thisRowCourses.append(thisCourse)
            self.Courses.append(thisRowCourses)

    # self.Courses（不定长的课程类列表）->self.Courses_text（5×7的文本列表）
    def _convertClass2Text(self):
        """
        self.Courses（不定长的课程类列表）->self.Courses_text（5×7的文本列表）
        """
        for row in self.Courses:  # 取出每一行的多门课程的信息
            thisRow = row.copy()  # 这一行的课程数据，是Course类的一维列表
            if len(thisRow) < 7:  # 如果这一行的课程数量不足7门，则先用None在最后补全（说明这一周有空课的，还是蛮好的）
                thisRow = thisRow + [None] * (7 - len(thisRow))
                # 倒序检查，如果有的课的位置和Column不对应，则将其换到对应的位置
                for i in range(6, -1, -1):
                    if thisRow[i] is not None and thisRow[i].column != i + 1:  # 此时它应去的位置是column-1
                        # 以下三行其实就是swap函数
                        temp = thisRow[thisRow[i].column - 1]
                        thisRow[thisRow[i].column - 1] = thisRow[i]
                        thisRow[i] = temp
            # 向列表中添加课程信息
            self.Courses_text.append(
                [x.HWCalenderMessage() if x is not None else "" for x in thisRow])

    # 考虑到不同软件对于课程表的格式的不同，这里允许不同子类对self.Courses_text进行修正，以便于更好地写入文件
    def polishText(self):
        """
        考虑到不同软件对于课程表的格式的不同，这里允许不同子类对self.Courses_text进行修正，以便于更好地写入文件
        """
        pass

    # 导出课程表到文件，由于有的文件csv，有的xlsx，所以需要指定文件类型
    def output(self, fileType, charset, autoOpen=True):
        """
        导出课程表到文件，由于有的文件csv，有的xlsx，所以需要指定文件类型
        """
        if "." not in fileType:
            fileType = "." + fileType
        with open(os_path.join(self.saveDir, self.AppName + "." + fileType), "w", encoding=charset, newline="") as csvfile:
            writer = csv_writer(csvfile)
            for row in self.Courses_text:
                writer.writerow(row)

        if autoOpen:
            os_startfile(os_path.join(
                self.saveDir, self.AppName + "." + fileType))


class HWCalender(Apps):
    """
    华为课程表的操作类，用于实现最终的转换、导出等功能
    """

    def __init__(self, sourceFile, time_1Week_Mon, saveDir):
        super().__init__(AppName="HWCalender", sourcePath=sourceFile,
                         time_1Week_Mon=time_1Week_Mon, saveDir=saveDir)
        self.polishText()

    def polishText(self):
        appendMessage = [["", "", "上课时间", "课程时长\n（分钟）"],
                         ["", "第1节", "8:30", "45"],
                         ["", "第2节", "9:25", "45"],
                         ["", "第3节", "10:30", "45"],
                         ["", "第4节", "11:25", "45"],
                         ["", "第5节", "14:00", "45"],
                         ["", "第6节", "14:55", "45"],
                         ["", "第7节", "16:00", "45"],
                         ["", "第8节", "16:55", "45"],
                         ["", "第9节", "19:00", "45"],
                         ["", "第10节", "19:55", "45"]]
        if len(self.Courses_text) < len(appendMessage):  # 如果课程信息比附加信息少，则补全课程信息的行数
            for i in range(len(self.Courses_text), len(appendMessage)):
                self.Courses_text.append(["", "", "", "", "", "", ""])
        for i in range(len(appendMessage)):  # 添加附加信息
            self.Courses_text[i] = self.Courses_text[i] + appendMessage[i]
        # 添加标题行
        self.Courses_text.insert(0,
                                 ["周一", "周二", "周三", "周四", "周五", "周六", "周日", "", "开学第一周周一日期", "", self.time_1Week_Mon])


def get1WeekMondayDate(Apps: str):
    content = input("请输入该学期第一周周一的日期，格式为'YYYY.MM.DD'：")
    if Apps == "HWCalender":
        if re_match(r"^\d{4}\.\d{2}\.\d{2}$", content):
            return content
        else:
            print("您的输入有误，请重试")
            return get1WeekMondayDate(Apps)


def askopenfilename(title, fileType):
    """
    请求文件链接
    """
    filePath = input(title)
    if '"' in filePath:  # 考虑到右键文件复制的链接中自带双引号，这里去掉
        filePath = filePath.replace('"', '')

    if not (os_path.isfile(path=filePath) and filePath.endswith(fileType)):
        print("您的输入有误，请重试")
        askopenfilename(title, fileType)
    else:
        return filePath


def askdirectory(title):
    dirPath = input(title)
    if '"' in dirPath:
        dirPath = dirPath.replace('"', '')
    if not os_path.isdir(dirPath):
        print("您的输入有误，请重试")
        askdirectory(title)
    else:
        return dirPath
