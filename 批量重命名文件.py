import os  # 重命名文件
import re
import openpyxl  # 操作excel


class Stu:
    """
    目前支持的信息有：姓名(name)、学号(id)，以后可以自行添加
    """

    def __init__(self, stu_name=None, stu_id=None):
        self.name = stu_name if stu_name else ''
        self.id = stu_id if stu_id else ''

    def show(self):
        print("姓名：{}，学号：{}".format(self.name, self.id))

    def message(self, pattern):
        if re.match('(姓名)(.)(学号)', pattern):
            return self.name + re.match('(姓名)(.)(学号)', pattern).group(2) + self.id
        elif re.match('(学号)(.)(姓名)', pattern):
            return self.id + re.match('(学号)(.)(姓名)', pattern).group(2) + self.name


class Commander:
    def __init__(self):
        self.defaultAllStuXlsxPath = r"E:\OneDrive\班级\20智科-名单.xlsx"
        self.allStuXlsxPath = input("请选择全部成员名单 (输入0使用默认名单):")  # 全部成员名单的路径
        self.allStuXlsxPath = self.defaultAllStuXlsxPath if self.allStuXlsxPath == '0' else self.allStuXlsxPath
        self.allStuSheet = openpyxl.load_workbook(
            self.allStuXlsxPath).active  # 全部成员名单的sheet
        self.nameDict = dict()  # 由姓名作为Key的字典
        self.netNameDict = dict()  # 由昵称作为Key的字典
        self.column_name = eval(input("请指定姓名所在的列 (输入0使用默认列):"))
        self.column_name = 1 if self.column_name == 0 else self.column_name
        self.column_id = eval(input("请指定学号所在的列 (输入0使用默认列):"))
        self.column_id = 2 if self.column_id == 0 else self.column_id
        self.readSheet()

    def readSheet(self, sheet=None, column_name=None, column_id=None):
        """
        读取sheet中的信息，并将信息存入allStu列表中
        :param sheet: sheet
        :param column_name: 姓名所在的列
        :param column_id: 学号所在的列
        """
        sheet = self.allStuSheet if sheet is None else sheet  # 使该方法具有通用性
        column_name = self.column_name if column_name is None else column_name
        column_id = self.column_id if column_id is None else column_id
        for i in range(2, self.allStuSheet.max_row + 1):  #
            stu = Stu()
            stu.name = sheet.cell(row=i, column=column_name).value
            stu.id = sheet.cell(row=i, column=column_id).value
            self.nameDict[stu.name] = stu
        print("全部成员名单已读取完毕，当前共有{}名成员".format(len(self.nameDict)))
        self.menu()

    def menu(self):
        print("1. 显示所有成员信息")
        print("2. 重命名")
        print("3. 补充昵称成员对应")
        print("4. 查看未出现的成员")
        print("5. 退出")
        choice = eval(input("请选择:"))
        if choice == 1:
            self.showAllStu()
        elif choice == 2:
            self.rename()
        elif choice == 3:
            self.addNetNameStu()
        elif choice == 4:
            self.findNotExist()
        elif choice == 5:
            exit(0)
        else:
            print("输入错误，请重新输入")
            self.menu()

    def showAllStu(self):
        for stu in self.nameDict:
            print("--------------------------------")
            self.nameDict[stu].show()
        print("共有{}名成员".format(len(self.nameDict)))
        self.menu()

    def addNetNameStu(self):
        """
        为了解决根据昵称找人的问题
        """
        appendXlsx = input("请选择昵称对应表")
        appendSheet = openpyxl.load_workbook(appendXlsx).active
        column_account = eval(input("请指定昵称所在的列:"))
        column_name = eval(input("请指定姓名所在的列:"))
        for i in range(2, appendSheet.max_row + 1):
            account = appendSheet.cell(row=i, column=column_account).value
            name = appendSheet.cell(row=i, column=column_name).value
            self.netNameDict[account] = self.nameDict[name]
        print("昵称对应完成")
        self.menu()

    def rename(self):
        renameDir = input("请选择重命名文件的目录")
        oldNamePattern = input("请输入原始名字的格式，如“昵称-学号”")
        newNamePattern = input("请输入新文件名格式，如“姓名-学号”:")
        for item in os.listdir(renameDir):
            if os.path.isfile(os.path.join(renameDir, item)):
                fileType = '.' + item.split('.')[-1]  # 获取文件扩展名
                account = None  # 获取用户对象
                try:
                    if re.match("(昵称)(.)(学号)", oldNamePattern):
                        account = self.netNameDict[re.split('[-.]', item)[0]]
                    elif re.match("(姓名)(.)(学号)", oldNamePattern):
                        account = self.nameDict[re.split('[-.]', item)[0]]
                    elif re.match("昵称", oldNamePattern):
                        account = self.netNameDict[re.split('[-.]', item)[0]]
                    elif re.match("姓名", oldNamePattern):
                        account = self.nameDict[re.split('[.]', item)[0]]
                    newName = account.message(pattern=newNamePattern)  # 获取新名字
                    os.rename(os.path.join(renameDir, item), os.path.join(
                        renameDir, newName + fileType))  # 执行重命名
                except KeyError:
                    print("获取文件对应失败，可能的失败原因如下：")
                    print("1.该同学信息未存在于全员信息表中，请检查", self.allStuXlsxPath, "的内容")
                    print("2.原有文件的命名与文件命名格式有冲突，请检查", item, "的命名格式")
        print('重命名完成')
        self.menu()

    def findNotExist(self):
        checkDir = input("请输入检查文件的目录:")
        namePattern = input("请输入文件名格式，如“姓名-学号”:")
        notExistStu = self.netNameDict.copy() if '昵称' in namePattern else self.nameDict.copy()
        for item in os.listdir(checkDir):
            if os.path.isfile(os.path.join(checkDir, item)):
                try:
                    if re.match("(昵称)(.)(学号)", namePattern):
                        del notExistStu[re.split('[-.]', item)[0]]
                    elif re.match("(姓名)(.)(学号)", namePattern):
                        del notExistStu[re.split('[-.]', item)[0]]
                    elif re.match("昵称", namePattern):
                        del notExistStu[re.split('[.]', item)[0]]
                    elif re.match("姓名", namePattern):
                        del notExistStu[re.split('[.]', item)[0]]
                except KeyError:
                    print("获取文件对应失败，可能的失败原因如下：")
                    print("1.该同学信息未存在于全员信息表中，请检查", self.allStuXlsxPath, "的内容")
                    print("2.原有文件的命名与文件命名格式有冲突，请检查",
                          os.path.join(checkDir, item), "的命名格式")
        print("未存在的成员如下:")
        for item in notExistStu:
            print("\t", item)
        print("共有", len(notExistStu), "位不在名单中，执行完毕")
        self.menu()


if __name__ == '__main__':
    commander = Commander()
    commander.menu()
