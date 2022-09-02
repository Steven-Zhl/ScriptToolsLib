from enum import Enum

import openpyxl


class Games(Enum):
    原神 = ["原神", "Genshin Impact", "GenshinImpact", "Genshin", ]
    战双帕弥什 = ["战双帕弥什", "Punishing: Gray Raven", "Punishing", "Gray Raven", ]
    明日方舟 = ["明日方舟", "Arknights", ]
    碧蓝航线 = ["碧蓝航线", "Azur Lane", "AzurLane", "Azur", "azur", "AZUR", ]
    崩坏3 = ["崩坏3", "Honkai", "崩坏学园", "崩壊3", "崩壊", ]
    尼尔机械纪元 = ["NieR:Automata", "NieR", "尼尔:自动人形",
              "automata", "Automata", "nier", ]
    最终幻想7 = ["最终幻想7", "FF7", "FF", ]
    艾尔登法环 = ["艾尔登法环", "Elden Ring", "ELDENRING", ]
    少女前线 = ["少女前线", "Girls' Frontline", "Frontline", ]
    无期迷途 = ["无期迷途", ]
    碧蓝档案 = ["碧蓝档案", "Blue Archive",  "BLUEARC", ]
    绝区零 = ["绝区零", "Zenless", "Zone Zero", ]
    幻塔 = ["幻塔", "Tower of Fantasy", ]
    鸣潮 = ["鸣潮", "鸣潮"]
    地下城与勇士 = ["地下城与勇士", "DNF", ]


class Sheet:
    """
    被操作的Sheet
    """

    def __init__(self, sheet: openpyxl.Workbook.worksheets, tags_col=3, row_length=9):
        """

        :param sheet:被操作的sheet
        :param tags_col: 划分游戏的内容所在列
        :param row_length: 每行内容的长度
        """
        self.sheet = sheet  # 被操作的工作表
        self.tags_col = tags_col  # 要检查的那一列
        self.row_length = row_length  # 每一行的内容区间

    def removeGameToNewXlsx(self, newXlsxPath: str, existTitleBar: bool, remain=False):
        """
        将当前（旧）文件中的内容根据游戏不同复制到新文件中

        :param newXlsxPath: 新文件的绝对路径
        :param existTitleBar: 是否存在标题栏
        :param remain: 把图片信息复制到新文件中时，是否保留原本文件中的图片信息
        """
        game_name_list = [item.name for item in Games]  # 获取所有的游戏名称
        # 为所有游戏创建Sheet，并初始化
        wb_new = openpyxl.Workbook()
        del wb_new["Sheet"]
        for name in game_name_list:
            wb_new.create_sheet(title=name)
            if existTitleBar:
                wb_new[name].append(
                    [cell.value for cell in tuple(self.sheet)[0]])

        # 逐行检测游戏
        for row_cells in (tuple(self.sheet) if not existTitleBar else tuple(self.sheet)[1:]):
            origin_row = self.Row(row=row_cells, row_length=self.row_length)
            for game in tuple(Games):
                if origin_row.checkGame(tags_col=self.tags_col, gameTags=game.value):
                    wb_new[game.name].append(
                        [cell.value for cell in origin_row.row])  # 将本行内容添加到新的文件中
                    if not remain:
                        origin_row.clear()
        wb_new.save(newXlsxPath)

    class Row:
        def __init__(self, row: tuple, row_length):
            self.row = row
            self.row_length = row_length

        def clear(self):
            """
            清理掉本行的内容
            """
            for item in self.row:
                item.value = ""

        def checkGame(self, tags_col, gameTags: list):
            """
            检查该行内容是否属于该游戏

            :param tags_col: tag所在的col
            :param gameTags: 用于硬匹配的tags
            :return: bool
            """
            tagContent = self.row[tags_col - 1].value
            for tag in gameTags:
                if tag in tagContent:
                    return True
                else:
                    continue
            else:
                return False


wb = openpyxl.load_workbook(filename="这里填写原文件路径")
ws = wb["这里填写被操作的sheet名称"]  # 如果只有一个sheet，这里可以改成ws = wb.active
# tags_col是检查tag的所在列；row_length是每行内容的长度
operate = Sheet(sheet=ws, tags_col=3, row_length=3)
operate.removeGameToNewXlsx(newXlsxPath="这里填写要保存的路径", existTitleBar=True,
                            remain=True)  # existTitleBar:是否有标题栏；remain:在复制时是否保留原有的内容
wb.save("这里填写原文件路径")
